import pandas as pd
from gmft.auto import CroppedTable, AutoTableDetector, AutoTableFormatter
from gmft.pdf_bindings import PyPDFium2Document
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Initialize table detector and formatter
detector = AutoTableDetector()
formatter = AutoTableFormatter()

def ingest_pdf(pdf_path):
    """ Extract tables from PDF file and return list of tables """
    doc = PyPDFium2Document(pdf_path)
    tables = []
    pages = []

    for page in doc:
        pages.append(page)
        detected_tables = detector.extract(page)
        if detected_tables:
            tables.extend(detected_tables)
    
    return tables, doc, pages

# Extract tables
pdf_path = "../ESG_reports/apple.pdf"
output_dir = "../labeled_files/apple_excel"
os.makedirs(output_dir, exist_ok=True)

tables, doc, pages = ingest_pdf(pdf_path)

# Create a new Excel workbook
wb = Workbook()
wb.remove(wb.active)  # Remove the default sheet

# Process each table
for i, table in enumerate(tables):
    print(f"\n--- Table {i + 1} ---")
    
    if hasattr(table, 'page') and table.page is not None:
        try:
            formatted_table = formatter.format(table)
            df = formatted_table.df()
            
            # Print table content
            print(df)
            
            # Create a new sheet for each table
            sheet = wb.create_sheet(title=f"Table_{i+1}")
            
            # Write the DataFrame to the sheet
            for r in dataframe_to_rows(df, index=False, header=True):
                sheet.append(r)
            
            # Add metadata
            sheet.cell(row=1, column=len(df.columns) + 2, value="Metadata")
            sheet.cell(row=2, column=len(df.columns) + 2, value="Number of Rows")
            sheet.cell(row=2, column=len(df.columns) + 3, value=len(df))
            sheet.cell(row=3, column=len(df.columns) + 2, value="Number of Columns")
            sheet.cell(row=3, column=len(df.columns) + 3, value=len(df.columns))
            
        except Exception as e:
            print(f"Error processing table {i + 1}: {e}")
    else:
        print(f"Table {i + 1} has no valid page reference.")

# Save the Excel file
excel_filename = os.path.join(output_dir, "all_tables.xlsx")
wb.save(excel_filename)
print(f"All tables saved to {excel_filename}")

# Close the document after processing all tables
doc.close()